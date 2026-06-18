import io
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from app.repositories.prediction import PredictionRepository
from app.repositories.model import ModelMetadataRepository
from app.core.exceptions import ValidationError
from uuid import UUID
from datetime import datetime

class ReportingService:
    def __init__(self, db: Session):
        self.db = db
        self.predict_repo = PredictionRepository(db)
        self.model_repo = ModelMetadataRepository(db)

    def generate_customer_pdf_report(self, prediction_id: UUID) -> bytes:
        prediction = self.predict_repo.get_by_id(prediction_id)
        if not prediction:
            raise ValidationError("Prediction log not found.")

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=letter,
            rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36
        )
        
        styles = getSampleStyleSheet()
        
        # Define Custom Styles for corporate aesthetic
        title_style = ParagraphStyle(
            'CorpTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=22,
            leading=26,
            textColor=colors.HexColor('#2563EB'), # Deep Blue
            spaceAfter=15
        )
        
        section_heading = ParagraphStyle(
            'CorpSection',
            parent=styles['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=14,
            leading=18,
            textColor=colors.HexColor('#1F2937'),
            spaceBefore=12,
            spaceAfter=8
        )
        
        body_style = ParagraphStyle(
            'CorpBody',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            leading=14,
            textColor=colors.HexColor('#374151')
        )

        story = []
        
        # Title Header
        story.append(Paragraph("PredictWise AI Churn Risk Audit", title_style))
        story.append(Paragraph(f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Log ID: {str(prediction.id)[:12]}", body_style))
        story.append(Spacer(1, 15))

        # Risk Summary Card Table
        risk_color = '#EF4444' if prediction.risk_category == 'High' else '#F59E0B' if prediction.risk_category == 'Medium' else '#10B981'
        summary_data = [
            [
                Paragraph("<b>Customer ID</b>", body_style), 
                Paragraph(prediction.customer_id, body_style)
            ],
            [
                Paragraph("<b>Risk Level</b>", body_style), 
                Paragraph(f"<font color='{risk_color}'><b>{prediction.risk_category} Risk</b></font>", body_style)
            ],
            [
                Paragraph("<b>Churn Probability</b>", body_style), 
                Paragraph(f"<b>{prediction.churn_probability:.2%}</b>", body_style)
            ]
        ]
        
        summary_table = Table(summary_data, colWidths=[150, 350])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')),
            ('PADDING', (0,0), (-1,-1), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        
        story.append(Paragraph("Churn Risk Assessment Summary", section_heading))
        story.append(summary_table)
        story.append(Spacer(1, 15))

        # Customer Details grid
        details_data = [
            [
                Paragraph("<b>Age</b>", body_style), Paragraph(str(prediction.age), body_style),
                Paragraph("<b>Gender</b>", body_style), Paragraph(prediction.gender, body_style)
            ],
            [
                Paragraph("<b>Location</b>", body_style), Paragraph(prediction.location, body_style),
                Paragraph("<b>Subscription</b>", body_style), Paragraph(prediction.subscription_type, body_style)
            ],
            [
                Paragraph("<b>Contract Duration</b>", body_style), Paragraph(f"{prediction.contract_duration} Mos", body_style),
                Paragraph("<b>Monthly Charges</b>", body_style), Paragraph(f"${prediction.monthly_charges:.2f}", body_style)
            ],
            [
                Paragraph("<b>Engagement Score</b>", body_style), Paragraph(f"{prediction.customer_engagement}/5", body_style),
                Paragraph("<b>Support Tickets</b>", body_style), Paragraph(str(prediction.support_tickets), body_style)
            ]
        ]
        
        details_table = Table(details_data, colWidths=[120, 130, 120, 130])
        details_table.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
            ('PADDING', (0,0), (-1,-1), 6),
            ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#F9FAFB')),
            ('BACKGROUND', (2,0), (2,-1), colors.HexColor('#F9FAFB')),
        ]))
        
        story.append(Paragraph("Account Characteristics", section_heading))
        story.append(details_table)
        story.append(Spacer(1, 15))

        # Risk Factors (SHAP)
        story.append(Paragraph("Primary Churn Risk Factors", section_heading))
        if prediction.explanation_json:
            shap_list = []
            for item in prediction.explanation_json[:3]:
                sign = "+" if item["impact"] > 0 else ""
                direction = "increases risk" if item["impact"] > 0 else "reduces risk"
                shap_list.append([
                    Paragraph(f"• <b>{item['feature'].replace('_', ' ')}</b> (Value: {item['value']})", body_style),
                    Paragraph(f"<font color='{risk_color if item['impact'] > 0 else '#10B981'}'><b>{sign}{item['impact']:.2f}</b></font> ({direction})", body_style)
                ])
            shap_table = Table(shap_list, colWidths=[250, 250])
            shap_table.setStyle(TableStyle([
                ('PADDING', (0,0), (-1,-1), 4),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ]))
            story.append(shap_table)
        else:
            story.append(Paragraph("No explainability parameters tracked.", body_style))
        story.append(Spacer(1, 15))

        # Retention Recommendations
        story.append(Paragraph("AI-Generated Retention Playbook", section_heading))
        if prediction.retention_suggestions:
            rec_list = []
            for s in prediction.retention_suggestions:
                rec_list.append([
                    Paragraph("★", ParagraphStyle('Star', parent=body_style, textColor=colors.HexColor('#F59E0B'), fontSize=12)),
                    Paragraph(s, body_style)
                ])
            rec_table = Table(rec_list, colWidths=[20, 480])
            rec_table.setStyle(TableStyle([
                ('PADDING', (0,0), (-1,-1), 4),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ]))
            story.append(rec_table)
        else:
            story.append(Paragraph("Standard communication loops recommended.", body_style))

        # Build PDF
        doc.build(story)
        pdf_bytes = buffer.getvalue()
        buffer.close()
        return pdf_bytes

    def generate_business_excel_report(self) -> bytes:
        predictions = self.predict_repo.list_all(limit=100000)
        summary = self.predict_repo.get_summary_metrics()
        
        wb = Workbook()
        
        # Sheet 1: Executive Dashboard
        ws1 = wb.active
        ws1.title = "Executive Summary"
        ws1.views.sheetView[0].showGridLines = True
        
        # Styles
        font_title = Font(name="Arial", size=16, bold=True, color="2563EB")
        font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        font_sub = Font(name="Arial", size=11, bold=True, color="1F2937")
        font_normal = Font(name="Arial", size=10)
        font_bold = Font(name="Arial", size=10, bold=True)
        
        fill_header = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        
        border_thin = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )
        
        # Title Block
        ws1["A1"] = "PredictWise AI Business Churn Analysis Report"
        ws1["A1"].font = font_title
        ws1["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws1["A2"].font = Font(name="Arial", size=9, italic=True)
        
        # Summary KPI Cards
        ws1["A4"] = "Platform Summary Metrics"
        ws1["A4"].font = font_sub
        
        kpis = [
            ("Total Scored Profiles", summary["total_customers"]),
            ("High Churn Risk Customers", summary["high_risk_customers"]),
            ("Average Churn Probability", f"{summary['average_churn_probability']:.2%}")
        ]
        
        for i, (kpi_name, kpi_val) in enumerate(kpis):
            row = 5 + i
            ws1.cell(row=row, column=1, value=kpi_name).font = font_bold
            ws1.cell(row=row, column=1).border = border_thin
            ws1.cell(row=row, column=2, value=kpi_val).font = font_normal
            ws1.cell(row=row, column=2).border = border_thin
            
        # Sheet 2: Scored Customer Records Catalog
        ws2 = wb.create_sheet(title="Customer Scoring Logs")
        ws2.views.sheetView[0].showGridLines = True
        
        headers = [
            "Timestamp", "Customer ID", "Age", "Gender", "Location", "Subscription",
            "Contract Duration", "Monthly Charges", "Total Charges", "Engagement",
            "Support Tickets", "Billing Status", "Churn Probability", "Risk Category"
        ]
        
        for col_idx, h in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center")
            
        for row_idx, pred in enumerate(predictions, 2):
            is_zebra = (row_idx % 2 == 0)
            row_data = [
                pred.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                pred.customer_id,
                pred.age,
                pred.gender,
                pred.location,
                pred.subscription_type,
                pred.contract_duration,
                pred.monthly_charges,
                pred.total_charges,
                pred.customer_engagement,
                pred.support_tickets,
                pred.payment_history,
                pred.churn_probability,
                pred.risk_category
            ]
            
            for col_idx, val in enumerate(row_data, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                cell.font = font_normal
                cell.border = border_thin
                
                # Percentage styling for churn probability
                if col_idx == 13:
                    cell.number_format = '0.0%'
                # Currency styling
                elif col_idx in [8, 9]:
                    cell.number_format = '$#,##0.00'
                    
                if is_zebra:
                    cell.fill = fill_zebra
                    
        # Adjust Column Widths on logs page
        for col in ws2.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)

        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()
        return excel_bytes
